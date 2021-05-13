# Import Module
# Code By Sumukha S
import tabula
from openpyxl import Workbook
import numpy as np
import csv
# read_pdf(PDF File Path, pages = Number of pages, **agrs)
# Read PDF File
# this contain a list
print('The input file (pdf) file(No extensions required)')
inpdfF=input()
inpdfF=inpdfF+".pdf"
print('The input file page numbers')
Pages=input()
Pages=int(Pages)
df = tabula.read_pdf(inpdfF, pages =Pages)[0]


# Convert into Excel File
print('Give the Path to download the sheet')
path=input()
print('File name')
name=input()
full=path+'\\'+name
full=full
df.to_csv(full, index = False)

# CSV to XLS
wb = Workbook()
ws = wb.active
with open(full) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=",")
    for row in csv_reader:
        ws.append(row)
    # All CSV to XLS
# d = ws.cell(row=3, column=7, value=)
wb.save(full+".xlsx")
j=0
for o in range(3,17):
    a=15 #append 15 cells space
    ws.cell(a+o,5).value = ws.cell(o, 5).value
for o in range(3,17):
    a = 15
    if ws.cell(a+o,5).value:
        j=j+1
        continue
    else:
        ws.delete_rows(a + o)
        print("row removed " + str(int(a + o)))
j=0
for o in range(18,30):
    if ws.cell(o,5).value:
        j=j+1
print("Number of subjects were "+str(j))
_cell = ws.cell(row=18 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=19 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=12 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=21 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=22 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=23 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=24 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=25 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=26 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=27 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=28 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=29 , column=5)
_cell.number_format = '0.00E+00'
_cell = ws.cell(row=30 , column=5)
_cell.number_format = '0.00E+00'
# for sub in range(18,30):
#     if int(ws.cell(sub,5).value)>=0 and int(ws.cell(sub,5).value)<=9:
#         ws.cell(sub,6).value=1
#     elif int(ws.cell(sub, 5).value)>=10 and int(ws.cell(sub, 5).value) <= 19:
#         ws.cell(sub, 6).value = 2
#     elif ws.cell(sub,5).value>=20 and ws.cell(sub,5).value<=29:
#         ws.cell(sub, 6).value = 3
#     elif ws.cell(sub, 5).value >=30 and ws.cell(sub, 5).value <= 39:
#         ws.cell(sub, 6).value = 4
#     elif ws.cell(sub, 5).value >=40 and ws.cell(sub, 5).value <= 49:
#         ws.cell(sub, 6).value = 5
#     elif ws.cell(sub, 5).value >=50 and ws.cell(sub, 5).value <= 59:
#         ws.cell(sub, 6).value = 6
#     elif ws.cell(sub,5).value>=60 and ws.cell(sub,5).value<=69:
#         ws.cell(sub, 6).value = 7
#     elif ws.cell(sub, 5).value >=70 and ws.cell(sub, 5).value <= 79:
#         ws.cell(sub, 6).value = 8
#     elif ws.cell(sub, 5).value >="80" and ws.cell(sub, 5).value <= "89":
#         ws.cell(sub, 6).value = 9
#     elif ws.cell(sub, 5).value >="90" and ws.cell(sub, 5).value <= "99":
#         ws.cell(sub, 6).value = 10
#     else:
#         ws.cell(sub,6).value="Invalid"
# Converting to number
if ws.cell(18, 5).value:
    st=ws.cell(18, 5).value
    ws['F18'] = float(st)
if ws.cell(19, 5).value:
    st=ws.cell(19, 5).value
    ws['F19'] = float(st)
if ws.cell(20, 5).value:
    st=ws.cell(20, 5).value
    ws['F20'] = float(st)
if ws.cell(21, 5).value:
    st=ws.cell(21, 5).value
    ws['F21'] = float(st)
if ws.cell(22, 5).value:
    st=ws.cell(22, 5).value
    ws['F22'] = float(st)
if ws.cell(23, 5).value:
    st=ws.cell(23, 5).value
    ws['F23'] = float(st)
if ws.cell(24, 5).value:
    st=ws.cell(24, 5).value
    ws['F24'] = float(st)
if ws.cell(25, 5).value:
    st=ws.cell(25, 5).value
    ws['F25'] = float(st)
if ws.cell(26, 5).value:
    st=ws.cell(26, 5).value
    ws['F26'] = float(st)
if ws.cell(27, 5).value:
    st=ws.cell(27, 5).value
    ws['F27'] = float(st)
if ws.cell(28, 5).value:
    st=ws.cell(28, 5).value
    ws['F28'] = float(st)
if ws.cell(29, 5).value:
    st=ws.cell(29, 5).value
    ws['F29'] = float(st)
if ws.cell(30, 5).value:
    st=ws.cell(30, 5).value
    ws['F30'] = float(st)






# Formulae





if ws.cell(18, 6).value:
    st=(ws.cell(18, 6).value/10)+1
    ws['G18'] = int(st)
if ws.cell(19, 6).value:
    st=(ws.cell(19, 6).value/10)+1
    ws['G19'] = int(st)
if ws.cell(20, 6).value:
    st=(ws.cell(20, 6).value/10)+1
    ws['G20'] = int(st)
if ws.cell(21, 6).value:
    st=(ws.cell(21, 6).value/10)+1
    ws['G21'] = int(st)
if ws.cell(22, 6).value:
    st=(ws.cell(22, 6).value/10)+1
    ws['G22'] = int(st)
if ws.cell(23, 6).value:
    st=(ws.cell(23, 6).value/10)+1
    ws['G23'] = int(st)
if ws.cell(24, 6).value:
    st=(ws.cell(24, 6).value/10)+1
    ws['G24'] = int(st)
if ws.cell(25, 6).value:
    st=(ws.cell(25, 6).value/10)+1
    ws['G25'] = int(st)
if ws.cell(26, 6).value:
    st=(ws.cell(26, 6).value/10)+1
    ws['G26'] = int(st)
if ws.cell(27, 6).value:
    st=(ws.cell(27, 6).value/10)+1
    ws['G27'] = int(st)
if ws.cell(28, 6).value:
    st=(ws.cell(28, 6).value/10)+1
    ws['G28'] = int(st)
if ws.cell(29, 6).value:
    st=(ws.cell(29, 6).value/10)+1
    ws['G29'] = int(st)
if ws.cell(30, 6).value:
    st=(ws.cell(30, 6).value/10)+1
    ws['G30'] = int(st)


# Input credits
print("Input credits for "+str(j)+" subjects Any EXTRA PUT 0")
sub=18
while sub<=(j+18):
    s=input()
    ws.cell(sub, 8).value=int(s)
    sub=sub+1

j=0
for o in range(18,30):
    if ws.cell(o,5).value:
        j=j+1
if ws.cell(18, 7).value:
    ws.cell(18, 9).value=float((ws.cell(18, 8).value)*(ws.cell(18, 7).value))
if ws.cell(19, 7).value:
    ws.cell(19, 9).value=float((ws.cell(19, 8).value)*(ws.cell(19, 7).value))
if ws.cell(20, 8).value:
    ws.cell(20, 9).value=float((ws.cell(20, 8).value)*(ws.cell(20, 7).value))
if ws.cell(21, 7).value:
    ws.cell(21, 9).value=float((ws.cell(21, 8).value)*(ws.cell(21, 7).value))
if ws.cell(22, 8).value:
    ws.cell(22, 9).value=float((ws.cell(22, 8).value)*(ws.cell(22, 7).value))
if ws.cell(23, 8).value:
    ws.cell(23, 9).value=float((ws.cell(23, 8).value)*(ws.cell(23, 7).value))
if ws.cell(24, 7).value:
    ws.cell(24, 9).value=float((ws.cell(24, 8).value)*(ws.cell(24, 7).value))
if ws.cell(25, 8).value:
    ws.cell(25, 9).value=float((ws.cell(25, 8).value)*(ws.cell(25, 7).value))
if ws.cell(26, 7).value:
    ws.cell(26, 9).value=float((ws.cell(26, 8).value)*(ws.cell(26, 7).value))
if ws.cell(27, 8).value:
    ws.cell(27, 9).value=float((ws.cell(27, 8).value)*(ws.cell(27, 7).value))
if ws.cell(28, 7).value:
    ws.cell(28, 9).value=float((ws.cell(28, 8).value)*(ws.cell(28, 7).value))
if ws.cell(29, 7).value:
    ws.cell(29, 9).value=float((ws.cell(29, 8).value)*(ws.cell(29, 7).value))
if ws.cell(30, 7).value:
    ws.cell(30, 9).value=float((ws.cell(30, 8).value)*(ws.cell(30, 7).value))
ws["H31"]="=SUM(H18:H30)"
ws["I31"]="=SUM(I18:I30)"
ws["J30"]="SGPA"
ws["J31"]="=I31/H31"


wb.save(full+"1"+".xlsx")



# C:\Users\User\Desktop\Upload
